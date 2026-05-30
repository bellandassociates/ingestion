# -*- Coding: utf-8 -*-
"""
test/conftest.py
Configs and fixtures for pytest
"""

# Built-in module imports
import io
import os
from pathlib import Path
import shutil

# 3rd party module imports
import openpyxl as xl
import psycopg
import pytest

DB_CONN_ERR_INCORRECT_TARGET_DB = """You might be trying to run tests against a
production database. Please check the target and try again."""

INCOMING_PATH = Path("/data")
SCHEMA_PATH = Path("/schema")

# -----------------------------------------------------------------------------
# Session-level environment checks
# -----------------------------------------------------------------------------

@pytest.fixture(scope="session", autouse=True)
def incoming_directory_exists():
    """
    Ensures that the watched directory actually exists at test initialization,
    and provides the path handle in an OS independent way.
    """
    assert INCOMING_PATH.exists(), (
        f"{INCOMING_PATH} does not exist in the container. "
        "Please ensure that the container is correctly configured."
    )

@pytest.fixture(scope="session", autouse=True)
def check_database_type():
    """
    Ensures that the test is not accidentally ran against a production level
    database by checking DB_TYPE environment variable.
    """
    assert os.environ["DB_TYPE"].lower() == "test", \
            DB_CONN_ERR_INCORRECT_TARGET_DB

# -----------------------------------------------------------------------------
# Per-test cleanup operations
# -----------------------------------------------------------------------------

@pytest.fixture(scope="function", autouse=True)
def purge_incoming_directory():
    """
    Ensures proper isolation between tests by erasing all side effects (files
    and database states) after each test.
    """
    # Clear incoming directory before test
    for item in INCOMING_PATH.iterdir():
        if item.is_file():
            item.unlink()
        elif item.is_dir():
            shutil.rmtree(item)
    yield

@pytest.fixture(scope="function", autouse=True)
def purge_test_database():
    """
    Purge the ephemeral test database to make sure the database is in a blank
    slate state for the test.
    """
    db_user = os.environ["DB_USER"]
    db_password = os.environ["DB_PASS"]
    host = os.environ["HOSTNAME"]

    admin_conn = psycopg.connect(
        f"postgresql://{db_user}:{db_password}@{host}:5432/postgres",
        autocommit=True,
    )
    with admin_conn.cursor() as cursor:
        cursor.execute("""
            SELECT datname
            FROM pg_database
            WHERE datistemplate = false
                AND datname NOT IN ('postgres');
        """)
        dbs = [row[0] for row in cursor.fetchall()]
        for db in dbs:
            cursor.execute(f"""
                SELECT pg_terminate_backend(pid)
                FROM pg_stat_activity
                WHERE datname = %s
                    AND pid <> pg_backend_pid();
            """,
            (db,)
            )
            cursor.execute(f"DROP DATABASE IF EXISTS {db}")

    admin_conn.close()
    yield

# -----------------------------------------------------------------------------
# Helper stubs
# -----------------------------------------------------------------------------

@pytest.fixture
def launch_ingestion_engine():
    """
    Launches ingestion engine. Currently a stub that does nothing.
    """
    def _launch():
        # TODO replace with real function call
        return None
    return _launch

@pytest.fixture
def write_test_file_to():
    """
    Creates a real test file to incoming path to simulate user file upload.
    """
    def _write(path, sheets):
        wb = xl.Workbook()
        wb.remove(wb.active)

        for sheet_name, rows in sheets.items():
            ws = wb.create_sheet(sheet_name)
            for row in rows:
                ws.append(row)
        wb.save(path)

    return _write

@pytest.fixture
def database_exist():
    """
    Stub for checking whether the database exists. As stub it always returns
    False.
    """
    def _exists(name):
        return False
    return _exists

@pytest.fixture
def tables_exist():
    """
    Stub for checking whether tables exist in the database. As stub it always
    returns False.
    """
    def _exists(table_names):
        return False
    return _exists

@pytest.fixture
def db():
    """
    Stub database inspector.
    """
    class StubDB:
        def columns_match(self, table, expected_columns, expected_types):
            return False

    return StubDB()

# -----------------------------------------------------------------------------
# Dummy fixtures
# -----------------------------------------------------------------------------

@pytest.fixture
def dummy_workbook():
    wb = xl.Workbook()
    wb.create_sheet("page1")
    wb.create_sheet("page2")
    contents = [
        ["row1", "val1a", "val1b"],
        ["row2", "val2a", "val2b"],
        ["row3", "val3a", "val3b"],
        ["row4", "val4a", "val4b"],
        [],
        ["col1", "col2", "col3", "col4"],
        ["val1a", "val2a", "val3a", "val4a"],
        ["val1b", "val2b", "val3b", "val4b"],
        ["val1c", "val2c", "val3c", "val4c"],
        ["val1d", "val2d", "val3d", "val4d"],
    ]
    for ws in wb:
        for row in contents:
            ws.append(row)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

# EOF
